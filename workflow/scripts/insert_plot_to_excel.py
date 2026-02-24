from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(snakemake.input["xlsx"])

for sample in wb.sheetnames:

    ws = wb[sample]
    ws.insert_rows(1, 23)

    img = Image(Path(snakemake.input["plot_dir"]) / (sample + ".png"))
    img.width = 1200
    img.height = 400

    ws.add_image(img, 'A2')

wb.save(snakemake.output["xlsx"])
